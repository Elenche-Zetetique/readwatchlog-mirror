#!/usr/bin/python3

from collections import defaultdict
from contextlib import contextmanager
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from rwl_base import BaseProcessor
from tqdm import tqdm
from typing import Any, Dict, Generator, List, Optional
from utilities import Utilities

class XlsxProcessor(BaseProcessor):
	"""
	A class to perform actions on a specially crafted XLSX (Excel) file.

	This class provides functionality to interact with Excel files, extract and process data,
	and interact with the YouTube Data API to retrieve video details. It includes methods for
	handling workbook operations, validating data, and managing outputs.

	Attributes:
		__slots__ (tuple): A tuple of instance variable names to optimize memory usage.

	Methods:
		_check_for_duplicates() -> dict
			Checks for duplicate links in an XLSX file and returns their indices.

		__check_record(row) -> bool
			Checks if a worksheet row represents an incomplete or invalid record.

		_convert_to_json() -> dict
			Converts the active worksheet into a JSON-compatible dictionary.

		__find_starting_row() -> int | None
			Finds the starting row containing a valid YouTube link.

		__get_col_number(col_name: str) -> int | None
			Finds the column number for a given column name in the worksheet.

		_get_links() -> dict
			Extracts YouTube links from an Excel worksheet and processes them to store video durations.

		_get_routines() -> dict
			Fetches daily routines from a worksheet as a JSON-structured dictionary.

		_get_tags() -> list[int]
			Extracts column numbers of cells containing 'Tag' in the first row.

		_order_tags() -> None
			Sorts tag values in worksheet rows in ascending order.

		_process_yt_link(link: str, row: int) -> dict
			Processes a YouTube link to extract video details and duration.
	"""
	__slots__ = ("_ws")

	def _check_for_duplicates(self) -> Dict[str, List[int]]:
		"""
		Checks for duplicate links in an XLSX file and returns their indices.

		Scans the worksheet starting at row 2, column 2, collecting all link values until an
		empty cell is encountered. Identifies duplicates and maps each duplicate link to a
		list of its zero-based indices in the original sequence.

		Returns:
			dict: Mapping of duplicate links to their indices, e.g., 
				{'link1': [0, 3], 'link2': [1, 5]}. Empty dict if no duplicates found.
		"""
		# Collect links with indices in one pass
		column = 2
		row = 2
		link_indices = defaultdict(list)

		while (value := self._ws.cell(row=row, column=column).value):
			link_indices[value].append(row - 2)  # Zero-based index
			row += 1

		# Filter to duplicates only
		return {
			link: indices
			for link, indices in link_indices.items()
			if len(indices) > 1
		}

	def __check_record(self, row) -> bool:
		"""
		Checks if a worksheet row represents an incomplete or invalid record.

		Examines the specified row in the worksheet (`self._ws`), retrieving values from the 'Duration',
		'Published', 'Author', and 'Exist' columns. Returns True if any of 'Duration', 'Published', or
		'Author' is a placeholder ('.') while 'Exist' is not a placeholder, indicating an incomplete record.

		Args:
			row (int): The row number in the worksheet to check.

		Returns:
			bool: True if the record is incomplete (has '.' in key fields but not in 'Exist'), False otherwise.

		Raises:
			AttributeError: If `self._ws` or `self.__get_col_number` is not initialized.
			ValueError: If `row` is not a valid integer or out of the worksheet's bounds.
			KeyError: If any required column ('Duration', 'Published', 'Author', 'Exist') is not found by
					`self.__get_col_number`.

		Notes:
			- Relies on `self._ws` as the worksheet object and `self.__get_col_number` to map column names
			to indices.
			- A '.' value indicates a placeholder or missing data in the respective field.
		"""
		# Cache column indices (assuming they don’t change per call)
		cols = {
			'Duration':	 self.__get_col_number('Duration'),
			'Published': self.__get_col_number('Published'),
			'Author':	 self.__get_col_number('Author'),
			'Exist':	 self.__get_col_number('Exist')
		}

		# Early exit if 'Exist' is not a placeholder
		if self._ws.cell(row=row, column=cols['Exist']).value != '.':
			return False

		# Check key fields for placeholders
		return any(
			self._ws.cell(row=row, column=cols[field]).value == '.'
			for field in ('Duration', 'Published', 'Author')
		)

	def __get_col_number(self, col_name: str) -> int | None:
		"""
		Finds the column number for a given column name in the worksheet.

		Searches row 1 of the worksheet for a cell matching the provided column name,
		returning its 1-based column number. Returns None if no match is found.

		Args:
			col_name: String representing the column name to search for.

		Returns:
			int: 1-based column number if the name is found, None otherwise.
		"""
		row = 1
		column = 1
		while (value := self._ws.cell(row=row, column=column).value):
			
			if value == col_name: return column
			column += 1

	def __get_last_row_number(self) -> int:
		"""
		Finds the last populated row number in the first column of the worksheet.

		Iterates through the rows of the first column (column 1) until an empty cell is encountered.
		The row number of the last non-empty cell is returned.

		Returns:
			int: The row number of the last populated cell in the first column.
		"""
		row = 1
		while self._ws.cell(row=row, column=1).value:
			row += 1
		return row

	def _convert_to_json(self) -> Dict[str, Dict[str, Any]]:
		"""
		Converts the active worksheet into a JSON-compatible dictionary.

		Uses the first row as column headers and the second column as unique keys. Each row’s
		remaining columns are mapped to nested key-value pairs. Dates in the 'Date' column,
		if present, are formatted as strings.

		Returns:
			dict: A dictionary where keys are second-column values and values are dictionaries of row data.
		
		Notes:
			- Assumes `_ws` is the worksheet object and the first row contains headers.
			- Removes the second column’s value from nested dictionaries.
			- Requires `datetime` for date formatting if 'Date' column exists.
		"""
		columns = []
		row, column = 1, 1

		# Extract column headers
		while (cell_value := self._ws.cell(row=row, column=column).value):
			columns.append(cell_value)
			column += 1

		vault = {}
		row = 2  # Start from the second row, assuming the first row is headers

		while (key := self._ws.cell(row=row, column=2).value):  # Use column 2 as keys
			row_data = {
				col: self._ws.cell(row=row, column=columns.index(col) + 1).value
				for col in columns
			}

			# Convert Date field to string if it exists and is a datetime object
			if "Date" in row_data and isinstance(row_data["Date"], datetime):
				row_data["Date"] = row_data["Date"].strftime("%d/%m/%y")

			vault[key] = row_data
			vault[key].pop(columns[1], None)  # Remove the second column's value from the dictionary
			row += 1

		return vault

	def __find_starting_row(self) -> int | None:
		"""
		Finds the starting row containing a valid YouTube link.

		Iterates through rows in a fixed column, searching for a cell containing 
		a YouTube link (`self._YT_PREFIX`). It returns the row index of the first 
		valid record that meets `self.__check_record(row)`.

		Returns:
			int: The row index of the first valid YouTube link.
		"""
		column = 2
		row = 2

		while link := self._ws.cell(row=row, column=column).value:
			if isinstance(link, str) and self._YT_PREFIX in link and self.__check_record(row=row):
				return row
			row += 1

	def _get_links(self) -> dict:
		"""
		Extracts and processes YouTube links from the worksheet, storing video durations.

		Iterates over rows in the specified range, identifies valid YouTube links, processes them,
		and updates the worksheet with extracted attributes (e.g., duration). Returns a dictionary
		of links and their metadata.

		Returns:
			dict: Mapping of YouTube links to their processed attributes.

		Raises:
			ValueError: If the range is invalid (e.g., `_END` <= `_START`) or `_START` is undefined with `_CHUNK`.

		Notes:
			- Uses `_START` and `_END` for row range, adjusted by `_CHUNK` or autosearch (`_AUTOSEARCH`).
			- Assumes `_ws` is the worksheet object and `_YT_PREFIX` defines valid YouTube link prefixes.
			- Caches column indices in `attr_columns` for efficiency.
			- Requires `tqdm` for progress tracking and private methods (`__find_starting_row`, `__get_last_row_number`, etc.).
		"""
		links: Dict[str, Dict[str, Any]] = {}
		LINK_COLUMN: int = 2  # Constant for link column index
		
		if self._AUTOSEARCH:
			self._START = self.__find_starting_row()
			self._END = self.__get_last_row_number()
			self._utilities.logger.info(f"Starting row: {self._START}")

		# Determine processing range efficiently
		if self._CHUNK:
			if self._START is None:
				raise ValueError("Value not found: --start is not defined")
			self._END = self._START + self._CHUNK

		# Pre-calculate column numbers to avoid repeated calls
		attr_columns: Dict[str, int] = {}
		
		# Process rows with progress tracking
		for row in tqdm(range(self._START, self._END), desc="Processing YouTube links"):
			link = self._ws.cell(row=row, column=LINK_COLUMN).value
			
			# Early continue for invalid links
			if not (isinstance(link, str) and self._YT_PREFIX in link):
				continue
				
			if not self.__check_record(row=row):
				continue
				
			# Process link and update worksheet
			link_info = self._process_yt_link(link=link, row=row)
			if not link_info or link not in link_info:
				continue
				
			# Cache column numbers and update cells efficiently
			for attribute, value in link_info[link].items():
				if value:
					if attribute not in attr_columns:
						attr_columns[attribute] = self.__get_col_number(col_name=attribute)
					col = attr_columns[attribute]
					if self._ws.cell(row=row, column=col).value == '.':
						self._ws.cell(row=row, column=col).value = value
			
			links.update(link_info)
		return links

	def _get_routines(self) -> Dict[str, Dict[str, float]]:
		"""
		Fetches daily routines from a worksheet as a JSON-structured dictionary.

		Extracts routine data starting from a predefined row (`self._START`), using column numbers
		for 'Date' and 'Duration'. For each row with a valid duration (not '.'), computes a rounded
		value, assigns a color based on cell fill, and aggregates totals by date and color.

		Returns:
			dict: JSON-structured dictionary with daily routines, formatted as:
				{
					"dd-mm-yyyy": {
						"green":	float,	# Rounded total for green routines
						"red":		float,	# Rounded total for red routines (optional)
						"yellow":	float	# Rounded total for yellow routines (optional)
					},
					...
				}
				Dates are strings in "dd-mm-yyyy" format; colors and values vary by row data.
		"""
		date_column = self.__get_col_number(col_name='Date')
		duration_column = self.__get_col_number(col_name='Duration')

		routines: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))  # Default to 0.0 for colors
		row = self._START

		while (date := self._ws.cell(row=row, column=date_column).value):
			
			if (duration_cell := self._ws.cell(row=row, column=duration_column)).value != '.':
				value = float(duration_cell.value)
				color = self._COLORS[duration_cell.fill.start_color.index]
				routines[date.strftime("%d-%m-%Y")][color] += value  # Accumulate directly

			row += 1

		# Convert to regular dict with rounded values
		return {
			date: {color: round(total, 2) for color, total in day.items()}
			for date, day in routines.items()
		}

	def _get_tags(self) -> List[int]:
		"""
		Extracts column numbers of cells containing 'Tag' in the first row.

		Scans row 1 of the worksheet, collecting 1-based column indices where the cell value
		contains the substring 'Tag'. Stops at the first empty cell.

		Returns:
			list[int]: List of 1-based column numbers where 'Tag' appears in the value.
		"""
		return [
			column
			for column, cell in enumerate(self._ws[1], start=1)
			if cell.value and 'Tag' in str(cell.value)
		]

	def _order_tags(self) -> None:
		"""
		Sorts tag values in worksheet rows in ascending order.

		Iterates over rows in the worksheet starting from `self._START`, collects non-placeholder ('.')
		tag values from columns specified in `self._get_tags()`, sorts them, and rewrites them across
		the same columns. Stops when the first tag column in a row has no value.

		Returns:
			None: Modifies the worksheet (`self._ws`) in place.

		Raises:
			AttributeError: If `self._ws`, `self._START`, or `self._get_tags()` is not properly initialized.
			TypeError: If tag column indices returned by `self._get_tags()` are not integers.

		Notes:
			- Relies on `self._ws` as the worksheet object and `self._START` as the starting row index.
			- Uses `self._get_tags()` to retrieve the list of column indices for tag values.
			- Placeholder values ('.') are ignored during sorting and not rewritten.
		"""
		tags: List[int] = self._get_tags()
		row = 2

		while self._ws.cell(row=row, column=tags[0]).value:

			# Collect and sort non-placeholder tag values
			tag_values = [
				cell.value for tag in tags
				if (cell := self._ws.cell(row=row, column=tag)).value != '.'
			]
			tag_values.sort()

			# Write sorted values back to tag columns
			for tag_col, value in zip(tags, tag_values):
				self._ws.cell(row=row, column=tag_col).value = value

			row += 1

	@contextmanager
	def _workbook_manager(self) -> Generator[Workbook, None, None]:
		"""
		A context manager that handles loading and saving an Excel workbook.

		Loads an Excel workbook from `self._FILE`, selects the worksheet specified by
		`self._SHEETNAME`, and yields the workbook object for use within a `with` block.
		Ensures the workbook is saved after execution, even if an error occurs, either to
		`self._FILE` or to a generated output file if `self._OUTPUT` is True.

		Yields:
			openpyxl.workbook.Workbook: The loaded Excel workbook object.

		Raises:
			TypeError: If `self._FILE` or `self._SHEETNAME` is not a string or is not set.
			FileNotFoundError: If the file specified by `self._FILE` does not exist.
			KeyError: If `self._SHEETNAME` does not exist in the workbook.
			Exception: For other unexpected errors during workbook loading or saving.

		Notes:
			- Relies on instance variables `self._FILE` (str), `self._SHEETNAME` (str),
			`self._OUTPUT` (bool), and `self._CUSTOM_NAME` (str or None).
			- Uses `self._utilities.generate_output_name` to create the output filename if
			`self._OUTPUT` is True.
			- The workbook is only saved if it was successfully loaded.
		"""
		wb: Optional[Workbook] = None
		try:
			wb = load_workbook(filename=self._FILE)
			self._ws = wb[self._SHEETNAME]
			yield wb  # Yield control to the wrapped function
		finally:
			if wb is not None:  # Ensure wb exists before trying to save
				if self._OUTPUT:
					filename = f"{self._utilities.generate_output_name(custom_name=self._CUSTOM_NAME)}.xlsx"
					wb.save(filename)
				else:
					filename = f"{self._utilities.generate_output_name(custom_name=self._ONLY_FILENAME)}.xlsx"
					wb.save(filename)